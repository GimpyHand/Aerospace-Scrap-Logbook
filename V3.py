import sys
import os
import shutil
import warnings
import sqlite3
import threading
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from PyQt5.QtWidgets import (QApplication, QWidget, QMainWindow, QVBoxLayout, QHBoxLayout, QGridLayout, QLabel, QLineEdit, QPushButton, QMessageBox, QComboBox, QFrame, QTableWidget, QTableWidgetItem, QCompleter)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QCursor, QIcon
import pandas as pd

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Define source and destination directories
source_dir = r'X:\ENGINE SERVICES\02 Work Orders'
destination_dir = r'X:\ENGINE SERVICES\Scrap Log Files'
log_file_path = r'X:\ENGINE SERVICES\Scrap Log Files\processed_directories.log'
db_file_path = r'X:\AEROSPACE\Aerospace YWG Scrap Parts Logbook\scrap_logbook.db'

class ScrapLogbook(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Cadorath Aerospace Scrap Logbook")
        self.setGeometry(280, 156, 1360, 768)
        self.setWindowIcon(QIcon('Cadorath-Logo.png'))

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)

        self.form_layout = QGridLayout()
        self.hbox_form_layout = QHBoxLayout()
        self.hbox_form_layout.addStretch(1)
        self.hbox_form_layout.addLayout(self.form_layout)
        self.hbox_form_layout.addStretch(1)
        self.main_layout.addLayout(self.hbox_form_layout)

        self.table_frame = QFrame(self)
        self.table_layout = QVBoxLayout(self.table_frame)
        self.main_layout.addWidget(self.table_frame)

        self.status_bar = QFrame(self)
        self.status_layout = QHBoxLayout(self.status_bar)
        self.main_layout.addWidget(self.status_bar)

        self.setup_form_elements()
        self.setup_status_bar()
        self.setup_table()

        self.initialize_db()
        self.load_data()

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.load_data)
        self.timer.start(30000)

        self.ensure_destination_directory()
        self.create_table_if_not_exists()

        # Start background processing
        self.start_background_processing()

    def start_background_processing(self):
        thread = threading.Thread(target=self.search_and_copy_files, args=(source_dir, destination_dir, log_file_path))
        thread.start()

    def setup_form_elements(self):
        self.label_date = QLabel("*Date (YYYY-MM-DD):", self)
        self.label_date.setStyleSheet("color: red;")
        self.form_layout.addWidget(self.label_date, 0, 0)

        self.entry_date = QLineEdit(self)
        self.form_layout.addWidget(self.entry_date, 0, 1)

        self.btn_date = QPushButton("📅", self)
        self.btn_date.clicked.connect(self.insert_current_date)
        self.form_layout.addWidget(self.btn_date, 0, 2)

        self.label_wo = QLabel("W/O:", self)
        self.form_layout.addWidget(self.label_wo, 1, 0)

        self.entry_wo = QLineEdit(self)
        self.form_layout.addWidget(self.entry_wo, 1, 1)

        self.label_part_number = QLabel("Part Number:", self)
        self.form_layout.addWidget(self.label_part_number, 2, 0)

        self.entry_part_number = QLineEdit(self)
        self.form_layout.addWidget(self.entry_part_number, 2, 1)

        self.part_numbers = self.fetch_part_numbers()
        self.completer = QCompleter(self.part_numbers, self)
        self.entry_part_number.setCompleter(self.completer)
        self.entry_part_number.textChanged.connect(self.update_part_description)

        self.label_part_description = QLabel("Part Description:", self)
        self.form_layout.addWidget(self.label_part_description, 3, 0)

        self.entry_part_description = QLineEdit(self)
        self.form_layout.addWidget(self.entry_part_description, 3, 1)

        self.label_serial_number = QLabel("Serial Number:", self)
        self.form_layout.addWidget(self.label_serial_number, 4, 0)

        self.entry_serial_number = QLineEdit(self)
        self.form_layout.addWidget(self.entry_serial_number, 4, 1)

        self.label_initials = QLabel("*Initials:", self)
        self.label_initials.setStyleSheet("color: red;")
        self.form_layout.addWidget(self.label_initials, 5, 0)

        self.entry_initials = QLineEdit(self)
        self.form_layout.addWidget(self.entry_initials, 5, 1)

        self.label_source = QLabel("*Source:", self)
        self.label_source.setStyleSheet("color: red;")
        self.form_layout.addWidget(self.label_source, 6, 0)

        self.combo_source = QComboBox(self)
        self.combo_source.addItems(["WPG MRO", "LAF MRO"])
        self.form_layout.addWidget(self.combo_source, 6, 1)

        self.label_remarks = QLabel("Remarks:", self)
        self.form_layout.addWidget(self.label_remarks, 7, 0)

        self.entry_remarks = QLineEdit(self)
        self.form_layout.addWidget(self.entry_remarks, 7, 1)

        self.btn_submit = QPushButton("Submit", self)
        self.btn_submit.clicked.connect(self.submit_data)
        self.form_layout.addWidget(self.btn_submit, 8, 0)

        self.btn_clear = QPushButton("Clear", self)
        self.btn_clear.clicked.connect(self.clear_form)
        self.form_layout.addWidget(self.btn_clear, 8, 1)

        self.label_filter = QLabel("Filter by Source:", self)
        self.form_layout.addWidget(self.label_filter, 9, 0)

        self.combo_filter = QComboBox(self)
        self.combo_filter.addItems(["All", "WPG MRO", "LAF MRO", "WPG ES", "LAF ES"])
        self.combo_filter.currentIndexChanged.connect(self.filter_data)
        self.form_layout.addWidget(self.combo_filter, 9, 1)

    def setup_status_bar(self):
        self.lbl_record_count = QLabel("Total records: 0", self)
        self.status_layout.addWidget(self.lbl_record_count)
        self.status_layout.addStretch(1)
        self.lbl_email_link = QLabel('<a href="mailto:alex.jessup@cadorath.com?subject=Scrap%20Log%20Issue">Contact Support</a>', self)
        self.lbl_email_link.setOpenExternalLinks(True)
        self.status_layout.addWidget(self.lbl_email_link)

    def setup_table(self):
        self.table_widget = QTableWidget(self)
        self.table_layout.addWidget(self.table_widget)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setColumnWidth(0, 150)
        self.table_widget.setColumnWidth(1, 100)
        self.table_widget.setColumnWidth(2, 100)
        self.table_widget.setColumnWidth(3, 150)
        self.table_widget.setColumnWidth(4, 150)
        self.table_widget.setColumnWidth(5, 150)
        self.table_widget.setColumnWidth(6, 50)
        self.table_widget.setColumnWidth(7, 100)
        self.table_widget.setStyleSheet("alternate-background-color: #f0f0f0;")
        self.table_widget.setSortingEnabled(True)
        self.table_widget.horizontalHeader().sectionClicked.connect(self.handle_header_click)

    def handle_header_click(self, logicalIndex):
        self.sort_order = self.table_widget.horizontalHeader().sortIndicatorOrder()
        self.table_widget.sortItems(logicalIndex, self.sort_order)

    def initialize_db(self):
        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scrap_logbook (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT,
                    wo TEXT,
                    part_number TEXT,
                    part_description TEXT,
                    serial_number TEXT,
                    initials TEXT,
                    remarks TEXT,
                    source TEXT
                )
            ''')

    def fetch_part_numbers(self):
        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT DISTINCT part_number FROM scrap_logbook')
            return [row[0] for row in cursor.fetchall()]

    def fetch_part_description(self, part_number):
        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT part_description FROM scrap_logbook WHERE part_number = ?', (part_number,))
            result = cursor.fetchone()
            return result[0] if result else ""

    def update_part_description(self):
        part_number = self.entry_part_number.text().strip()
        part_description = self.fetch_part_description(part_number)
        self.entry_part_description.setText(part_description)

    def submit_data(self):
        date = self.entry_date.text().strip()
        initials = self.entry_initials.text().strip().upper()
        serial_number = self.entry_serial_number.text().strip().upper()
        source = self.combo_source.currentText()

        if not date or not initials or not source:
            missing_fields = []
            if not date:
                missing_fields.append("Date")
            if not initials:
                missing_fields.append("Initials")
            if not source:
                missing_fields.append("Source")
            QMessageBox.warning(self, "Missing Information", f"Please fill out the following required field(s): {', '.join(missing_fields)}")
            return

        try:
            datetime.strptime(date, '%Y-%m-%d')
        except ValueError:
            QMessageBox.warning(self, "Invalid Date", "Please enter the date in YYYY-MM-DD format.")
            return

        # Include the current time in the date field
        date_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        wo = self.entry_wo.text().strip()
        part_number = self.entry_part_number.text().strip()
        part_description = self.entry_part_description.text().strip()
        remarks = self.entry_remarks.text().strip()

        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO scrap_logbook (date, source, wo, part_number, part_description, serial_number, initials, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (date_time, source, wo, part_number, part_description, serial_number, initials, remarks))

        self.load_data()
        QMessageBox.information(self, "Success", "Data submitted successfully.")
        self.clear_form()

    def load_data(self):
        self.filter_data()

    def filter_data(self):
        selected_source = self.combo_filter.currentText()
        query = '''
            SELECT date, source, wo, part_number, part_description, serial_number, initials, remarks
            FROM scrap_logbook
        '''
        params = []
        if selected_source != "All":
            query += " WHERE source = ?"
            params.append(selected_source)
        query += " ORDER BY date DESC"

        with sqlite3.connect(db_file_path) as conn:
            df = pd.read_sql_query(query, conn, params=params)

        self.table_widget.setRowCount(len(df))
        self.table_widget.setColumnCount(len(df.columns))

        custom_headers = ["Date", "Source", "Work Order", "Part Number", "Description", "Serial No.", "Initials", "Remarks"]
        self.table_widget.setHorizontalHeaderLabels(custom_headers)

        for i, row in df.iterrows():
            for j, value in enumerate(row):
                if j == 0:  # Format the date to exclude the time
                    value = value.split(' ')[0]
                item = QTableWidgetItem(str(value))
                if j in [4, 7]:  # Indexes of "Description" and "Remarks" columns
                    item.setTextAlignment(Qt.AlignLeft)
                else:
                    item.setTextAlignment(Qt.AlignCenter)
                self.table_widget.setItem(i, j, item)

        self.table_widget.setColumnWidth(0, 80)
        self.table_widget.setColumnWidth(1, 80)
        self.table_widget.setColumnWidth(2, 80)
        self.table_widget.setColumnWidth(3, 125)
        self.table_widget.setColumnWidth(4, 300)
        self.table_widget.setColumnWidth(5, 120)
        self.table_widget.setColumnWidth(6, 50)
        self.table_widget.setColumnWidth(7, 430)
        self.table_widget.setWordWrap(True)
        self.update_record_count()

    def update_record_count(self):
        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM scrap_logbook')
            count = cursor.fetchone()[0]
        self.lbl_record_count.setText(f"Total records: {count}")

    def clear_form(self):
        self.entry_date.clear()
        self.entry_wo.clear()
        self.entry_part_number.clear()
        self.entry_part_description.clear()
        self.entry_serial_number.clear()
        self.entry_initials.clear()
        self.entry_remarks.clear()
        self.combo_source.setCurrentIndex(0)

    def insert_current_date(self):
        self.entry_date.setText(datetime.now().strftime('%Y-%m-%d'))

    def ensure_destination_directory(self):
        os.makedirs(destination_dir, exist_ok=True)

    def create_table_if_not_exists(self):
        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS scrap_logbook (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date TEXT,
                    wo TEXT,
                    part_number TEXT,
                    part_description TEXT,
                    serial_number TEXT,
                    initials TEXT,
                    remarks TEXT,
                    source TEXT
                )
            ''')

    def is_complete(self, file_path):
        try:
            workbook = load_workbook(file_path, data_only=True)
            if "WO Data" in workbook.sheetnames:
                sheet = workbook["WO Data"]
                return sheet["L2"].value and "complete" in sheet["L2"].value.lower()
        except (InvalidFileException, PermissionError, BadZipFile) as e:
            print(f"Error reading {file_path}: {e}")
        return False

    def extract_data(self, file_path):
        try:
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = ["Scrap Part List", "Unserviceable Part List", "Unservicable Part List"]
            data = []
            for sheet_name in sheet_names:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    wo = sheet["I3"].value
                    if wo:
                        wo = ''.join(filter(str.isdigit, str(wo)))  # Convert to string and extract only digits
                    for range_start in [12, 43, 74]:
                        for row in sheet.iter_rows(min_row=range_start, max_row=range_start+19, min_col=3, max_col=7):
                            part_description, part_number, serial_number, qty, remarks = [cell.value for cell in row]
                            if part_description and part_number and serial_number and qty:
                                remarks = f"{remarks} (Qty: {qty})" if remarks else f"Qty: {qty}"
                                data.append((wo, part_number, part_description, serial_number, remarks))
            return data
        except (InvalidFileException, PermissionError, BadZipFile) as e:
            print(f"Error reading {file_path}: {e}")
        return []

    def insert_data_into_db(self, data):
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with sqlite3.connect(db_file_path) as conn:
            cursor = conn.cursor()
            cursor.executemany('''
                INSERT INTO scrap_logbook (date, wo, part_number, part_description, serial_number, remarks, initials, source)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', [(current_date, *entry, 'ES', 'WPG ES') for entry in data])

    def search_and_copy_files(self, source_dir, destination_dir, log_file_path):
        processed_dirs = set()
        if os.path.exists(log_file_path):
            with open(log_file_path, 'r') as log_file:
                processed_dirs = set(log_file.read().splitlines())

        new_processed_dirs = set()

        for root, _, files in os.walk(source_dir):
            folder_name = os.path.basename(root)
            print(f"Scanning directory: {folder_name}")  # Print the directory being scanned
            if folder_name in processed_dirs:
                print(f"Skipping directory: {folder_name}")  # Print the directory being skipped
                continue  # Skip already processed directories

            copied = False
            for file in files:
                if file.endswith(('.xlsx', '.xlsm')) and not file.startswith('~$'):
                    file_path = os.path.join(root, file)
                    if self.is_complete(file_path):
                        new_file_name = f"{folder_name}_{file}"
                        destination_path = os.path.join(destination_dir, new_file_name)
                        shutil.copy(file_path, destination_path)
                        copied = True
                        # Extract data and insert into database
                        data = self.extract_data(file_path)
                        if data:
                            self.insert_data_into_db(data)

            if copied:
                new_processed_dirs.add(folder_name)

        with open(log_file_path, 'a') as log_file:
            for dir_name in new_processed_dirs:
                log_file.write(f"{dir_name}\n")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ScrapLogbook()
    window.show()
    sys.exit(app.exec_())