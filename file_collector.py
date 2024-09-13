import os
import shutil
import warnings
import sqlite3
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile

# Suppress specific warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Define source and destination directories
source_dir = r'X:\ENGINE SERVICES\02 Work Orders'
destination_dir = r'X:\ENGINE SERVICES\Scrap Log Files'
log_file_path = r'processed_directories.log'
db_file_path = r'scrap_logbook.db'

# Function to check if "complete" is in cell L2 of "WO Data" worksheet
def is_complete(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        if "WO Data" in workbook.sheetnames:
            sheet = workbook["WO Data"]
            return sheet["L2"].value and "complete" in sheet["L2"].value.lower()
    except (InvalidFileException, PermissionError, BadZipFile) as e:
        print(f"Error reading {file_path}: {e}")
    return False

# Function to extract data from specified worksheets and tables
def extract_data(file_path):
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet_names = ["Scrap Part List", "Unserviceable Part List", "Unservicable Part List"]
        data = []
        for sheet_name in sheet_names:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                wo = sheet["I3"].value
                if wo:
                    wo = ''.join(filter(str.isdigit, str(wo)))  # Convert to string and extract only digits
                for range_start in [12, 43, 74]:
                    for row in sheet.iter_rows(min_row=range_start, max_row=range_start+19, min_col=3, max_col=7):
                        part_description, part_number, serial_number, qty, remarks = [cell.value for cell in row]
                        if part_description and part_number and serial_number and qty:
                            remarks = f"{remarks} (Qty: {qty})" if remarks else f"Qty: {qty}"
                            data.append((wo, part_number, part_description, serial_number, remarks))
        return data
    except (InvalidFileException, PermissionError, BadZipFile) as e:
        print(f"Error reading {file_path}: {e}")
    return []

# Function to create the database table if it doesn't exist
def create_table_if_not_exists():
    conn = sqlite3.connect(db_file_path)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scrap_logbook (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            wo TEXT,
            part_number TEXT,
            part_description TEXT,
            serial_number TEXT,
            initials TEXT,
            remarks TEXT,
            source TEXT
        )
    ''')
    conn.commit()
    conn.close()

# Function to insert data into the database
def insert_data_into_db(data):
    conn = sqlite3.connect(db_file_path)
    cursor = conn.cursor()
    current_date = datetime.now().strftime('%Y-%m-%d')
    for entry in data:
        cursor.execute('''
            INSERT INTO scrap_logbook (date, wo, part_number, part_description, serial_number, remarks, initials)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (current_date, *entry, 'ES'))  # Provide "ES" as the default value for initials
    conn.commit()
    conn.close()

# Function to recursively search and copy files
def search_and_copy_files(source_dir, destination_dir, log_file_path):
    # Read processed directories from log file
    if os.path.exists(log_file_path):
        with open(log_file_path, 'r') as log_file:
            processed_dirs = set(log_file.read().splitlines())
    else:
        processed_dirs = set()

    new_processed_dirs = set()

    for root, _, files in os.walk(source_dir):
        folder_name = os.path.basename(root)
        print(f"Scanning directory: {folder_name}")  # Print the directory being scanned
        if folder_name in processed_dirs:
            print(f"Skipping directory: {folder_name}")  # Print the directory being skipped
            continue  # Skip already processed directories

        copied = False
        for file in files:
            if file.endswith(('.xlsx', '.xlsm')) and not file.startswith('~$'):
                file_path = os.path.join(root, file)
                if is_complete(file_path):
                    new_file_name = f"{folder_name}_{file}"
                    destination_path = os.path.join(destination_dir, new_file_name)
                    shutil.copy(file_path, destination_path)
                    copied = True
                    # Extract data and insert into database
                    data = extract_data(file_path)
                    if data:
                        insert_data_into_db(data)

        if copied:
            new_processed_dirs.add(folder_name)

    # Write updated processed directories to log file
    with open(log_file_path, 'a') as log_file:
        for dir_name in new_processed_dirs:
            log_file.write(f"{dir_name}\n")

# Ensure the destination directory exists
os.makedirs(destination_dir, exist_ok=True)

# Create the database table if it doesn't exist
create_table_if_not_exists()

# Start the search and copy process
search_and_copy_files(source_dir, destination_dir, log_file_path)